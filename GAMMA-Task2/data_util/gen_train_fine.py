import os

import cv2
import torch
import torchvision.transforms as transforms
from PIL import Image
from efficientnet_pytorch import EfficientNet
from openpyxl import load_workbook
from torch.autograd import Variable

# idx
idx = 4

# get gt
loc_map = dict()
wb = load_workbook('../dataset/train_loc_aug/training_GT.xlsx')
sheet = wb['Sheet1']
for i in range(1, sheet.max_row + 1):
    if i == 1:
        continue
    img = sheet.cell(row=i, column=1).value + '.jpg'
    w, h = sheet.cell(row=i, column=2).value, sheet.cell(row=i, column=3).value
    loc_map[img] = [w, h]

# gen train_img
fine_size = 512
train_fine = dict()
with open('../dataset/train_loc_aug/train_name_loc_{}.txt'.format(idx)) as input_file:
    lines = input_file.readlines()
for i in range(len(lines)):
    a1, a2, a3, a4, a5 = lines[i].strip().split()
    train_fine[a1] = [a4, a5]

img_map = dict()
img_path = '../dataset/train_loc_aug/img'
for img in os.listdir(img_path):
    im = Image.open(img_path + '/' + img)
    img_map[img] = im.size
    x, y = int(loc_map[img][0]), int(loc_map[img][1])
    im = cv2.imread(img_path + '/' + img)
    im = im[y - 256:y + 256, x - 256:x + 256]
    loc_map[img][0] = fine_size // 2
    loc_map[img][1] = fine_size // 2
    cv2.imwrite('../dataset/train_loc_fine/img_{0}/{1}'.format(idx, img), im)
    print(img)

print('gen train_img success')

# gen train_txt
f = open("../dataset/train_loc_fine/train_name_loc_{}.txt".format(idx), "w")
for k, v in train_fine.items():
    print(k, v)
    f.write("{0} {1} {2} {3} {4}\n".
            format(k, fine_size, fine_size, loc_map[k][0], loc_map[k][1]))

print('gen train_txt success')

# load model
ckpt_path = '../result' \
            '/2021.8.30/4_b4_15.09_epoch_1000_lr_0.0001_bsz_4_coarse' \
            '/best_model_667.pth'
model = EfficientNet.from_pretrained('efficientnet-b4', num_classes=2)
pretrained = torch.load(ckpt_path)
model.load_state_dict(pretrained)
model = model.cuda()
model.eval()

# gen test_txt,test_img
test_img_list = []
test_fine = dict()
with open('../dataset/train_loc_aug/test_name_loc_{}.txt'.format(idx)) as input_file:
    lines = input_file.readlines()
for i in range(len(lines)):
    a1, _, _, a4, a5 = lines[i].strip().split()
    test_img_list.append(a1)
    test_fine[a1] = [a4, a5]

mean = [0.37758928, 0.18647607, 0.05481077]
std = [0.26575511, 0.14306038, 0.06513966]
te_transform = transforms.Compose([
    transforms.Resize((224, 224)),
    transforms.ToTensor(),
    transforms.Normalize(mean, std)
])


def pil_loader(path):
    with open(path, 'rb') as f:
        with Image.open(f) as img:
            return img.convert('RGB')


f = open("../dataset/train_loc_fine/test_name_loc_{}.txt".format(idx), "w")
aug_img_size = 2000
max_loss = 0.0
for test_img in test_img_list:
    img_path = '../dataset/train_loc_aug/img/' + test_img
    img = pil_loader(img_path)
    img = te_transform(img)

    with torch.no_grad():
        img = Variable(img.cuda()).unsqueeze(0)
        out = model(img)
        pixel_size = torch.Tensor([aug_img_size, aug_img_size]).cuda()
        pixel_out = torch.mul((out + 1) / 2, pixel_size)

        im = cv2.imread(img_path)
        p_x, p_y = pixel_out[0][0].item(), pixel_out[0][1].item()
        x, y = int(p_x), int(p_y)
        print('predict:', test_img, fine_size // 2, fine_size // 2)
        im = im[y - 256:y + 256, x - 256:x + 256]
        cv2.imwrite('../dataset/train_loc_fine/img_{}/'.format(idx) + test_img, im)

        affine_x, affine_y = p_x - 256, p_y - 256
        real_x = float(test_fine[test_img][0]) - affine_x
        real_y = float(test_fine[test_img][1]) - affine_y
        print('GT:', test_img, real_x, real_y)
        f.write("{0} {1} {2} {3} {4}\n".
                format(test_img, fine_size, fine_size, real_x, real_y))

        pixel_out = torch.tensor([fine_size // 2, fine_size // 2]).cuda()
        pixel_label = torch.tensor([real_x, real_y]).cuda()
        new_loss = torch.dist(pixel_out, pixel_label).item()
        print('Loss:', test_img, new_loss)
        if new_loss > max_loss:
            max_loss = new_loss

print('gen train_img train_txt success')
print('max loss:', max_loss)
