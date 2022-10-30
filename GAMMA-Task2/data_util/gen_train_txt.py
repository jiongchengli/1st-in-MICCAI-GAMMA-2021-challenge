import os
import random
from PIL import Image
from openpyxl import load_workbook

target = 5

img_map = dict()
img_path = '../dataset/train_loc_aug/img'
for img in os.listdir(img_path):
    im = Image.open(img_path + '/' + img)
    img_map[img] = im.size
    # print('图片：%s, 宽：%d,高：%d' % (img, im.size[0], im.size[1]))


loc_map = dict()
wb = load_workbook('../dataset/train_loc_aug/training_GT.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
for i in range(1, sheet.max_row + 1):
    if i == 1:
        continue
    img = sheet.cell(row=i, column=1).value + '.jpg'
    w, h = sheet.cell(row=i, column=2).value, sheet.cell(row=i, column=3).value
    loc_map[img] = [w, h]
    # print(img, w, h)

print(loc_map)

data_idx = set([i for i in range(1, 101)])
print(data_idx)
# train_idx = set([i for i in range(1, 81)])
train_idx = set(random.sample(range(1, 101), 80))
print(train_idx)
test_idx = data_idx - train_idx
print(test_idx)


f = open("../dataset/train_loc_aug/train_name_loc_{}.txt".format(target), "w")
for idx in train_idx:
    format_idx = '%04d' % idx
    img = format_idx + '.jpg'
    img_size = img_map[img]
    img_loc = loc_map[img]
    print(img, img_loc)
    f.write("{0} {1} {2} {3} {4}\n".format(img, img_size[0], img_size[1],
                                           img_loc[0], img_loc[1]))


f = open("../dataset/train_loc_aug/test_name_loc_{}.txt".format(target), "w")
for idx in test_idx:
    format_idx = '%04d' % idx
    img = format_idx + '.jpg'
    img_size = img_map[img]
    img_loc = loc_map[img]
    f.write("{0} {1} {2} {3} {4}\n".format(img, img_size[0], img_size[1],
                                           img_loc[0], img_loc[1]))
