import os

import cv2
from PIL import Image
import torchvision.transforms as transforms
from openpyxl import load_workbook

loc_map = dict()
wb = load_workbook('./training_GT.xlsx')
sheet = wb['Sheet1']
for i in range(1, sheet.max_row + 1):
    if i == 1:
        continue
    img = sheet.cell(row=i, column=1).value + '.jpg'
    w, h = sheet.cell(row=i, column=2).value, sheet.cell(row=i, column=3).value
    loc_map[img] = [w, h]

img_map = dict()
img_path = '../dataset/val_loc/img'
for img in os.listdir(img_path):
    im = Image.open(img_path + '/' + img)
    img_map[img] = im.size
    im = im.convert('RGB')
    loc_map[img][0] += (2000 - im.size[0]) // 2
    loc_map[img][1] += (2000 - im.size[1]) // 2
    print(im.size[0], ' : ', im.size[1])
    im = transforms.CenterCrop((2000, 2000))(im)
    im.save("./val_loc_aug/{}".format(img))
    # new GT
    x, y = int(loc_map[img][0]), int(loc_map[img][1])
    im = cv2.imread("./img/{}".format(img))
    cv2.line(im, (x, y - 50), (x, y + 50), (0, 0, 0), 3)
    cv2.line(im, (x - 50, y), (x + 50, y), (0, 0, 0), 3)
    cv2.imwrite('./aug_GT/{0}'.format(img), im)


# save xlsx
for i in range(1, sheet.max_row + 1):
    if i == 1:
        continue
    img = sheet.cell(row=i, column=1).value + '.jpg'
    sheet.cell(i, 2).value = loc_map[img][0]
    sheet.cell(i, 3).value = loc_map[img][1]
wb.save('./training_GT.xlsx')
